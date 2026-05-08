"""
Excel upload service for bulk importing:
  - Attendance records
  - Internal marks
  - Student data
Supports both .xlsx and .csv formats.
"""

import csv
import io
import os
from datetime import datetime, date
from typing import Any

from models import db
from models.user import User
from models.subject import Subject
from models.attendance import Attendance
from models.internal_mark import InternalMark


class ExcelService:
    """Handles parsing and importing Excel/CSV data into the database."""

    # ── Attendance Import ─────────────────────────────────────

    @staticmethod
    def import_attendance(filepath: str, uploaded_by: int | None = None) -> dict[str, Any]:
        """
        Import attendance from Excel/CSV file.

        Expected columns: Student ID, Subject Code, Date (YYYY-MM-DD), Status (PRESENT/ABSENT/LATE)

        Returns summary dict with counts and errors.
        """
        rows = ExcelService._read_file(filepath)
        if not rows:
            return {'success': False, 'message': 'Could not read file or file is empty', 'imported': 0, 'errors': []}

        imported = 0
        skipped = 0
        errors = []

        for i, row in enumerate(rows, start=2):  # start=2 because row 1 is header
            try:
                student_id_str = str(row.get('Student ID', row.get('student_id', row.get('StudentID', '')))).strip()
                subject_code = str(row.get('Subject Code', row.get('subject_code', row.get('SubjectCode', '')))).strip()
                date_str = str(row.get('Date', row.get('date', ''))).strip()
                status = str(row.get('Status', row.get('status', 'PRESENT'))).strip().upper()

                if not all([student_id_str, subject_code, date_str]):
                    errors.append(f"Row {i}: Missing required fields (Student ID, Subject Code, or Date)")
                    continue

                if status not in ('PRESENT', 'ABSENT', 'LATE'):
                    status = 'PRESENT'

                # Lookup student
                student = User.query.filter_by(student_id=student_id_str, role='STUDENT').first()
                if not student:
                    errors.append(f"Row {i}: Student ID '{student_id_str}' not found")
                    continue

                # Lookup subject
                subject = Subject.query.filter_by(code=subject_code).first()
                if not subject:
                    errors.append(f"Row {i}: Subject code '{subject_code}' not found")
                    continue

                # Parse date
                try:
                    att_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                except ValueError:
                    try:
                        att_date = datetime.strptime(date_str, '%d/%m/%Y').date()
                    except ValueError:
                        errors.append(f"Row {i}: Invalid date format '{date_str}' (use YYYY-MM-DD)")
                        continue

                # Check for existing record
                existing = Attendance.query.filter_by(
                    student_id=student.id,
                    subject_id=subject.id,
                    date=att_date
                ).first()

                if existing:
                    existing.status = status
                    existing.is_manual = True
                    existing.marked_by = uploaded_by
                    existing.marked_at = datetime.utcnow()
                    skipped += 1
                else:
                    attendance = Attendance(
                        student_id=student.id,
                        subject_id=subject.id,
                        date=att_date,
                        time=datetime.now().time(),
                        status=status,
                        marked_by=uploaded_by,
                        is_manual=True,
                        notes='Imported from Excel'
                    )
                    db.session.add(attendance)
                    imported += 1

            except Exception as e:
                errors.append(f"Row {i}: {str(e)}")
                continue

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return {'success': False, 'message': f'Database error: {str(e)}', 'imported': 0, 'errors': errors}

        return {
            'success': True,
            'message': f'Successfully imported {imported} records, updated {skipped} existing records',
            'imported': imported,
            'updated': skipped,
            'errors': errors,
            'total_processed': imported + skipped + len(errors)
        }

    # ── Internal Marks Import ─────────────────────────────────

    @staticmethod
    def import_internal_marks(filepath: str, uploaded_by: int | None = None) -> dict[str, Any]:
        """
        Import internal marks from Excel/CSV file.

        Expected columns: Student ID, Subject Code, Exam Type, Marks, Max Marks

        Returns summary dict.
        """
        rows = ExcelService._read_file(filepath)
        if not rows:
            return {'success': False, 'message': 'Could not read file or file is empty', 'imported': 0, 'errors': []}

        imported = 0
        updated = 0
        errors = []

        for i, row in enumerate(rows, start=2):
            try:
                student_id_str = str(row.get('Student ID', row.get('student_id', row.get('StudentID', '')))).strip()
                subject_code = str(row.get('Subject Code', row.get('subject_code', row.get('SubjectCode', '')))).strip()
                exam_type = str(row.get('Exam Type', row.get('exam_type', row.get('ExamType', '')))).strip().upper()
                marks_str = str(row.get('Marks', row.get('marks', row.get('Marks Obtained', '')))).strip()
                max_marks_str = str(row.get('Max Marks', row.get('max_marks', row.get('MaxMarks', '50')))).strip()

                if not all([student_id_str, subject_code, exam_type, marks_str]):
                    errors.append(f"Row {i}: Missing required fields")
                    continue

                # Lookup student
                student = User.query.filter_by(student_id=student_id_str, role='STUDENT').first()
                if not student:
                    errors.append(f"Row {i}: Student ID '{student_id_str}' not found")
                    continue

                # Lookup subject
                subject = Subject.query.filter_by(code=subject_code).first()
                if not subject:
                    errors.append(f"Row {i}: Subject code '{subject_code}' not found")
                    continue

                marks = float(marks_str)
                max_marks = float(max_marks_str) if max_marks_str else 50.0

                # Check for existing record
                existing = InternalMark.query.filter_by(
                    student_id=student.id,
                    subject_id=subject.id,
                    exam_type=exam_type
                ).first()

                if existing:
                    existing.marks_obtained = marks
                    existing.max_marks = max_marks
                    existing.uploaded_by = uploaded_by
                    existing.uploaded_at = datetime.utcnow()
                    updated += 1
                else:
                    mark = InternalMark(
                        student_id=student.id,
                        subject_id=subject.id,
                        exam_type=exam_type,
                        marks_obtained=marks,
                        max_marks=max_marks,
                        semester=student.semester,
                        uploaded_by=uploaded_by
                    )
                    db.session.add(mark)
                    imported += 1

            except ValueError as e:
                errors.append(f"Row {i}: Invalid number format — {str(e)}")
                continue
            except Exception as e:
                errors.append(f"Row {i}: {str(e)}")
                continue

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return {'success': False, 'message': f'Database error: {str(e)}', 'imported': 0, 'errors': errors}

        return {
            'success': True,
            'message': f'Successfully imported {imported} marks, updated {updated} existing',
            'imported': imported,
            'updated': updated,
            'errors': errors,
            'total_processed': imported + updated + len(errors)
        }

    # ── Bulk Student Import ───────────────────────────────────

    @staticmethod
    def import_students(filepath: str) -> dict[str, Any]:
        """
        Import students from Excel/CSV file.

        Expected columns: Username, Full Name, Email, Student ID, Department, Semester, Section, Password (optional)

        Returns summary dict.
        """
        rows = ExcelService._read_file(filepath)
        if not rows:
            return {'success': False, 'message': 'Could not read file or file is empty', 'imported': 0, 'errors': []}

        imported = 0
        skipped = 0
        errors = []

        for i, row in enumerate(rows, start=2):
            try:
                username = str(row.get('Username', row.get('username', ''))).strip()
                full_name = str(row.get('Full Name', row.get('full_name', row.get('Name', '')))).strip()
                email = str(row.get('Email', row.get('email', ''))).strip()
                student_id_str = str(row.get('Student ID', row.get('student_id', row.get('StudentID', '')))).strip()
                department = str(row.get('Department', row.get('department', 'Computer Science'))).strip()
                semester_str = str(row.get('Semester', row.get('semester', ''))).strip()
                section = str(row.get('Section', row.get('section', ''))).strip()
                password = str(row.get('Password', row.get('password', 'student123'))).strip()

                if not all([username, full_name, email, student_id_str]):
                    errors.append(f"Row {i}: Missing required fields (Username, Full Name, Email, Student ID)")
                    continue

                # Check duplicates
                if User.query.filter_by(username=username).first():
                    errors.append(f"Row {i}: Username '{username}' already exists")
                    skipped += 1
                    continue

                if User.query.filter_by(email=email).first():
                    errors.append(f"Row {i}: Email '{email}' already exists")
                    skipped += 1
                    continue

                if User.query.filter_by(student_id=student_id_str).first():
                    errors.append(f"Row {i}: Student ID '{student_id_str}' already exists")
                    skipped += 1
                    continue

                semester = int(semester_str) if semester_str and semester_str.isdigit() else None

                user = User(
                    username=username,
                    email=email,
                    full_name=full_name,
                    role='STUDENT',
                    student_id=student_id_str,
                    department=department,
                    semester=semester,
                    section=section if section else None
                )
                user.set_password(password)
                db.session.add(user)
                imported += 1

            except Exception as e:
                errors.append(f"Row {i}: {str(e)}")
                continue

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return {'success': False, 'message': f'Database error: {str(e)}', 'imported': 0, 'errors': errors}

        return {
            'success': True,
            'message': f'Successfully imported {imported} students, skipped {skipped} duplicates',
            'imported': imported,
            'skipped': skipped,
            'errors': errors,
            'total_processed': imported + skipped + len(errors)
        }

    # ── File Reading Helpers ──────────────────────────────────

    @staticmethod
    def _read_file(filepath: str) -> list[dict]:
        """Read Excel or CSV file and return list of row dicts."""
        ext = os.path.splitext(filepath)[1].lower()

        if ext in ('.xlsx', '.xls'):
            return ExcelService._read_excel(filepath)
        elif ext == '.csv':
            return ExcelService._read_csv(filepath)
        else:
            return []

    @staticmethod
    def _read_excel(filepath: str) -> list[dict]:
        """Read .xlsx file using openpyxl."""
        try:
            from openpyxl import load_workbook  # pyre-ignore[21]
            wb = load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                return []

            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                return []

            headers = [str(h).strip() if h else f'col_{i}' for i, h in enumerate(rows[0])]
            result = []
            for row in rows[1:]:
                row_dict = {}
                for j, val in enumerate(row):
                    if j < len(headers):
                        row_dict[headers[j]] = val if val is not None else ''
                if any(v for v in row_dict.values() if str(v).strip()):
                    result.append(row_dict)
            wb.close()
            return result
        except Exception as e:
            print(f"[ExcelService] Error reading Excel file: {e}")
            return []

    @staticmethod
    def _read_csv(filepath: str) -> list[dict]:
        """Read CSV file."""
        try:
            with open(filepath, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                return list(reader)
        except Exception as e:
            print(f"[ExcelService] Error reading CSV file: {e}")
            return []
